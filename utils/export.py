import io
from typing import Union

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Column display metadata
# ---------------------------------------------------------------------------
COLUMN_META = {
    "title":         {"label": "Job Title",        "width": 32},
    "company":       {"label": "Company",           "width": 22},
    "location":      {"label": "Location",          "width": 22},
    "description":   {"label": "Description",       "width": 55},
    "date_posted":   {"label": "Date Posted",       "width": 16},
    "date_accessed": {"label": "Date Accessed",     "width": 20},
    "source":        {"label": "Source",            "width": 16},
    "link":          {"label": "Link",              "width": 45},
}

# Colours (hex without '#')
COLOR_HEADER_BG   = "1A2744"
COLOR_HEADER_FONT = "4ECDC4"
COLOR_ROW_ALT     = "EFF6FF"
COLOR_ROW_WHITE   = "FFFFFF"
COLOR_LINK        = "0563C1"
COLOR_BORDER      = "D0D7E3"


def export_to_excel(df: pd.DataFrame) -> bytes:
    """
    Export a job-results DataFrame to a formatted Excel workbook.

    Returns the workbook as raw bytes so it can be passed directly to
    st.download_button.
    """
    # Rename columns to friendly labels
    rename_map = {k: v["label"] for k, v in COLUMN_META.items() if k in df.columns}
    df_export = df.rename(columns=rename_map).copy()

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Job Results")
        ws = writer.sheets["Job Results"]

        # ------------------------------------------------------------------
        # Header row styling
        # ------------------------------------------------------------------
        header_fill = PatternFill(
            start_color=COLOR_HEADER_BG,
            end_color=COLOR_HEADER_BG,
            fill_type="solid",
        )
        header_font = Font(
            bold=True,
            color=COLOR_HEADER_FONT,
            size=11,
            name="Calibri",
        )
        thin_border = Border(
            bottom=Side(style="thin", color=COLOR_BORDER),
        )
        header_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        for col_idx in range(1, len(df_export.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        ws.row_dimensions[1].height = 28

        # ------------------------------------------------------------------
        # Column widths (map by friendly label)
        # ------------------------------------------------------------------
        label_to_width = {v["label"]: v["width"] for v in COLUMN_META.values()}
        for col_idx, col_name in enumerate(df_export.columns, start=1):
            col_letter = get_column_letter(col_idx)
            width = label_to_width.get(col_name, 18)
            ws.column_dimensions[col_letter].width = width

        # ------------------------------------------------------------------
        # Data rows – alternating fill + wrap text
        # ------------------------------------------------------------------
        alt_fill   = PatternFill(start_color=COLOR_ROW_ALT,   end_color=COLOR_ROW_ALT,   fill_type="solid")
        white_fill = PatternFill(start_color=COLOR_ROW_WHITE, end_color=COLOR_ROW_WHITE, fill_type="solid")
        cell_alignment = Alignment(vertical="top", wrap_text=True)

        for row_idx in range(2, len(df_export) + 2):
            fill = alt_fill if row_idx % 2 == 0 else white_fill
            for col_idx in range(1, len(df_export.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fill
                cell.alignment = cell_alignment
            ws.row_dimensions[row_idx].height = 42

        # ------------------------------------------------------------------
        # Make "Link" column clickable hyperlinks
        # ------------------------------------------------------------------
        link_col_idx: Union[int, None] = None
        for idx, name in enumerate(df_export.columns, start=1):
            if name == COLUMN_META["link"]["label"]:
                link_col_idx = idx
                break

        if link_col_idx is not None:
            link_font = Font(color=COLOR_LINK, underline="single", name="Calibri")
            for row_idx in range(2, len(df_export) + 2):
                cell = ws.cell(row=row_idx, column=link_col_idx)
                if cell.value:
                    cell.hyperlink = str(cell.value)
                    cell.font = link_font

        # ------------------------------------------------------------------
        # Freeze header + enable auto-filter
        # ------------------------------------------------------------------
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    return output.getvalue()
